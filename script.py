import xlrd
import MySQLdb
import openpyxl

# Open the workbook and define the worksheet
book = xlrd.open_workbook("dba.xlsx")
sheet = book.sheet_by_name("sho1")

# Establish a MySQL connection
database = MySQLdb.connect (host="localhost", user = "root", passwd = "password", db = "shomie", use_unicode=True, charset="utf8")

# Get the cursor, which is used to traverse the database, line by line
cursor = database.cursor()

# Create the INSERT INTO sql query
query = """INSERT INTO properties (description, adress, number, floor, type, price, capacity, availibility, number_wcs, latitude, longitude, has_living_room, has_cleaning, expenses_included, flatmates, has_terrace, only_girls, landlord_id, route, room_id, presentation, zone, stay) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

# Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
for r in range(1, sheet.nrows):
	  description    = sheet.cell(r,17).value
	  adress            = sheet.cell(r,2).value           
	  number            = sheet.cell(r,4).value
	  floor            = sheet.cell(r,3).value
	  type            = sheet.cell(r,5).value
	  price             = sheet.cell(r,7).value
	  capacity             = sheet.cell(r,11).value
	  availibility             = sheet.cell(r,18).value
	  number_wcs             = sheet.cell(r,10).value
	  latitude             = sheet.cell(r,19).value
	  longitude             = sheet.cell(r,20).value
	  has_living_room             = sheet.cell(r,12).value
	  has_cleaning             = sheet.cell(r,9).value
	  expenses_included             = sheet.cell(r,8).value
	  flatmates             = sheet.cell(r,21).value
	  has_terrace             = sheet.cell(r,13).value
	  only_girls             = sheet.cell(r,14).value
	  landlord_id             = sheet.cell(r,15).value
          route            = sheet.cell(r,16).value
          room_id    = sheet.cell(r,0).value
          presentation    = sheet.cell(r,22).value
          zone    = sheet.cell(r,23).value
          stay    = sheet.cell(r,24).value	
	  
     
	      

      # Assign values from each row
	  values = (description, adress, number, floor, type, price, capacity, availibility, number_wcs, latitude, longitude, has_living_room, has_cleaning, expenses_included, flatmates, has_terrace, only_girls, landlord_id, route, room_id, presentation, zone, stay)

	  # Execute sql Query
	  cursor.execute(query, values)
	  
# Close the cursor
cursor.close()

# Commit the transaction
database.commit()

# Close the database connection
database.close()

# Print results
print ("")
print ("All Done! Bye, for now.")
print ("")
columns = str(sheet.ncols)
rows = str(sheet.nrows)
# print ("I just imported " + %2B columns %2B + " columns and " + %2B rows %2B + " rows to MySQL!")